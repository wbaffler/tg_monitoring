import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image

class FormXlsx(object):
      

    def __init__(self, data_list):
        self.data_list = data_list
        self.header = ['Номер', 
                    'Канал', 
                    'Текст сообщения',
                    'Дата публикации', 
                    'Количество просмотров',
                    'Ссылка на публикацию', 
                    'Количество негативных реакций', 
                    'Количество нейтральных реакций', 
                    'Количество позитивных реакций', 
                    'Реакции']


    def __emoji_list_to_excel(self, data):
            for d in data:
                emoji_string = ''
                for emoji in d['emoji_list']:
                    emoji_string += f'{emoji["emoji"]} - {emoji["count"]}\n'
                d['emoji_list'] = emoji_string
            return data
            

    def form_xlsx_file(self, date):
            
            '''Формирование страницы с графиками'''
            
            workbook = Workbook()
            workbook.remove(workbook.active)
            sheet = workbook.create_sheet('Графики')
            img = Image(f'temp/visualisation.png')
            img.width = 1440
            img.height = 1000
            sheet.add_image(img, 'A1')
            
            '''Формирование страницы с таблицей публикаций'''
            table_sheet = workbook.create_sheet('Таблица публикаций')
            
            table_sheet.append(self.header)
            new_data_list = self.__emoji_list_to_excel(self.data_list)
            headers_dict = list(new_data_list[0].keys())
            print(new_data_list)
            for row in new_data_list:
                row_data = [row[header] for header in headers_dict]
                table_sheet.append(row_data)

            filename_path = f'docs/data{date}.xlsx'
            workbook.save(filename_path)

            return filename_path



